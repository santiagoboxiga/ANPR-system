import os
import re
import statistics
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

car_id_pattern = re.compile(r'car_([\d\.]+)_')

def extract_car_id(filename):
    match = car_id_pattern.search(filename)
    return match.group(1) if match else "Unknown"

def analyze_plate(alpr, image_path):
    results = alpr.predict(image_path)
    if not results or not results[0].ocr:
        return "Unknown", None

    ocr = results[0].ocr
    text = ocr.text
    conf = ocr.confidence

    if isinstance(conf, list):
        return text, statistics.mean(conf)
    return text, conf

def remove_duplicate_plates(data):
    seen = set()
    unique = []
    for record in data:
        if record[1] not in seen:
            unique.append(record)
            seen.add(record[1])
    return unique

def write_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Car ID", "Number Plate", "Image", "% of Confidence"])

    for row_idx, (car_id, plate_text, image_path, conf) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=car_id)
        ws.cell(row=row_idx, column=2, value=str(plate_text))

        img = XLImage(image_path)
        img.width = 100
        img.height = 50
        ws.add_image(img, f"C{row_idx}")

        ws.cell(row=row_idx, column=4, value=f"{conf * 100:.2f}%" if conf else "N/A")

    wb.save(output_file)
